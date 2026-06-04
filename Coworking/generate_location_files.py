"""
Generate one Markdown file per coworking location from coworking_location_directory_enriched.xlsx.

Output folder: Coworking/locations/
Naming:        {operator}_{market}_{address_slug}.md  (all lowercase)

Usage:
    python generate_location_files.py
    python generate_location_files.py --dry-run   # print filenames only

Requires: pip install openpyxl
"""

import argparse
import os
import re
from datetime import date
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

EXCEL_PATH   = Path(__file__).parent / "coworking_location_directory_enriched.xlsx"
OUTPUT_DIR   = Path(__file__).parent / "locations"
TODAY        = date.today().isoformat()

# Operator contacts -- keyed by (operator, market). "all" means any market.
CONTACTS = {
    ("WeWork",               "SF"):     ("Jack Ortlieb",     "jack.ortlieb@wework.com",           "(949) 903-9991", "Market Director, Broker Partnerships"),
    ("WeWork",               "NYC"):    ("Conor Golden",     "conor.golden@wework.com",            "(516) 592-1616", "Senior Leasing Director, Tri-State"),
    ("WeWork",               "Boston"): ("Garrett McCready", "garrett.mccready@wework.com",        "(617) 633-4568", "Leasing Director"),
    ("Industrious",          "all"):    ("Julissa Cajigas",  "jcajigas@industriousoffice.com",     "(917) 789-8731", "Broker Sales Lead"),
    ("IWG / Regus",          "all"):    ("Eric Fletcher",    "Portals.US@iwgplc.com",              "(619) 220-6021", "Broker Account Manager"),
    ("Regus",                "all"):    ("Eric Fletcher",    "Portals.US@iwgplc.com",              "(619) 220-6021", "Broker Account Manager"),
    ("IWG / Spaces",         "all"):    ("Sara Parker",      "mia.king.offices@gmail.com",         "(972) 764-8882", "Marketing"),
    ("Spaces",               "all"):    ("Sara Parker",      "mia.king.offices@gmail.com",         "(972) 764-8882", "Marketing"),
    ("Mindspace",            "all"):    ("Chris Chamoun",    "chris.c@mindspace.me",               "(202) 802-5246", "US Lead, Enterprise Sales & Broker Partnerships"),
    ("Tishman Speyer Studio","all"):    ("Josh Barton",      "jbarton@tishmanspeyer.com",          "(847) 702-6166", "Head of Sales"),
    ("Expansive",            "all"):    ("Carly Erickson",   "carly@expansive.com",                "(623) 253-7191", "Director of Sales, West"),
}


def get_contact(operator, market):
    key = (operator, market)
    if key in CONTACTS:
        return CONTACTS[key]
    key_all = (operator, "all")
    if key_all in CONTACTS:
        return CONTACTS[key_all]
    return ("", "", "", "")


def slugify(text):
    text = str(text).lower().strip()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[\s_/\\]+", "_", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("_")


def make_filename(operator, market, address):
    # Use first line of address only
    addr_line = str(address).split("\n")[0].strip()
    return f"{slugify(operator)}_{slugify(market)}_{slugify(addr_line)}.md"


def make_md(row):
    operator    = row.get("Operator", "") or row.get("operator", "")
    market      = row.get("Market", "") or row.get("market", "")
    name        = row.get("Location Name", "") or row.get("location_name", "")
    address     = str(row.get("Address", "") or row.get("address", "")).strip()
    website     = row.get("Website", "") or row.get("website", "") or ""
    tandem      = row.get("Tandem Listing", "") or row.get("tandem_listing", "") or ""
    price       = row.get("Starting Price/mo", "") or row.get("monthly_price_from", "") or ""
    amenities_raw = row.get("Amenities & Features", "") or row.get("amenities", "") or ""
    needs_review = str(row.get("Needs Review", "") or row.get("needs_review", "")).strip().lower()

    # Skip cookie-consent fallback text from failed page fetches
    amenities = amenities_raw if amenities_raw and "cookie" not in amenities_raw.lower()[:50] else ""

    contact_name, contact_email, contact_phone, contact_title = get_contact(operator, market)

    # YAML frontmatter
    addr_single = address.replace("\n", ", ").replace('"', "'")
    frontmatter = f"""---
operator: {operator}
market: {market}
location_name: {name}
address: "{addr_single}"
website: "{website}"
tandem_listing: "{tandem}"
monthly_price_from: "{price}"
last_updated: {TODAY}
contact_name: {contact_name}
contact_email: {contact_email}
contact_phone: "{contact_phone}"
contact_title: {contact_title}
needs_review: {needs_review == "true"}
---"""

    # Markdown body
    website_line  = f"[{website}]({website})" if website else "—"
    tandem_line   = f"[View on Tandem]({tandem})" if tandem else "*Not yet listed on Tandem*"
    price_line    = f"From ${price}/mo" if price else "Contact operator for pricing"

    body = f"""
# {operator} — {name} — {market}

## Overview

| Field | Value |
|-------|-------|
| Operator | {operator} |
| Market | {market} |
| Address | {addr_single} |
| Website | {website_line} |
| Tandem Listing | {tandem_line} |
| Starting Price | {price_line} |

## Points of Contact

| Field | Value |
|-------|-------|
| Name | {contact_name or "—"} |
| Title | {contact_title or "—"} |
| Email | {f"[{contact_email}](mailto:{contact_email})" if contact_email else "—"} |
| Phone | {contact_phone or "—"} |

## Building Features

{amenities if amenities else "*No amenity information on file.*"}

## Current Availability

*Last updated: {TODAY} — No availability on file. Contact operator for current options.*

## Outreach History

| Date | Direction | Notes |
|------|-----------|-------|
"""

    return frontmatter + body


def load_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {headers[i]: (row[i] if row[i] is not None else "") for i in range(len(headers))}
        rows.append(d)
    return rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    rows = load_excel(EXCEL_PATH)
    print(f"Loaded {len(rows)} rows from {EXCEL_PATH.name}")

    OUTPUT_DIR.mkdir(exist_ok=True)

    written = 0
    seen = set()
    for row in rows:
        op   = str(row.get("Operator", "") or row.get("operator", "")).strip()
        mkt  = str(row.get("Market", "") or row.get("market", "")).strip()
        addr = str(row.get("Address", "") or row.get("address", "")).strip()
        if not op or not addr:
            continue

        filename = make_filename(op, mkt, addr)

        # Deduplicate -- keep first occurrence
        if filename in seen:
            print(f"  SKIP (duplicate): {filename}")
            continue
        seen.add(filename)

        if args.dry_run:
            print(f"  {filename}")
        else:
            content = make_md(row)
            (OUTPUT_DIR / filename).write_text(content, encoding="utf-8")
            written += 1

    if args.dry_run:
        print(f"\nDry run -- {len(seen)} files would be written to {OUTPUT_DIR}")
    else:
        print(f"\nWrote {written} files to {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
